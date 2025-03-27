import openpyxl
from config import config
from module.box import read_boxes_from_config
import cv2
def write_to_excel(file_path, frame_result):
    """
    将数据和图像写入 Excel 文件，支持直接在表格中查看图像
    """
    import tempfile
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image as ExcelImage
    import os
    from datetime import datetime
    
    # 创建临时目录存储图片
    temp_dir = os.path.join(tempfile.gettempdir(), "train_images")
    os.makedirs(temp_dir, exist_ok=True)
    
    # 创建/加载工作簿
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
        ws = wb.active

        if is_same_with_last_row(ws, frame_result):
            return
    else:
        wb = Workbook()
        ws = wb.active
        # 创建固定表头
        headers = ["时间"]
        # 按配置顺序生成列名（与boxes.json配置顺序一致）
        for box in read_boxes_from_config(config.BOX_CONFIG_PATH):
            headers.append(f"{box.name}_图片")
            headers.append(f"{box.name}_结果")
        ws.append(headers)
    
    # 准备行数据
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    new_row = [timestamp]
    image_paths = {}
    
    # 处理每个识别结果
    for name, data in frame_result.items():
        # 保存结果文本
        new_row.append("")
        new_row.append(data.get("result_text", ""))
        
        # 保存图片到临时文件
        if data.get("image") is not None:
            img_path = os.path.join(temp_dir, f"{name}_{timestamp.replace(':', '-')}.png")
            cv2.imwrite(img_path, data["image"])
            img = ExcelImage(img_path)
            image_paths[name] = img
    
    # 先写入文本数据
    ws.append(new_row)
    
    # 插入图片到对应单元格
    last_row = ws.max_row
    from openpyxl.styles import Alignment  # 导入 Alignment 类
    for col_offset, (name, img) in enumerate(image_paths.items(), start=1):
        # 计算图片插入的列号(字母)
        col_letter = openpyxl.utils.get_column_letter(col_offset*2)
        cell = f"{col_letter}{last_row}"
        img.anchor = cell
        ws.add_image(img)
    
    # 调整行高和列宽
    ws.row_dimensions[last_row].height = 20  # 设置行高
    for col in range(1, ws.max_column + 1):  # 从第 1 列开始遍历所有列
        col_letter = openpyxl.utils.get_column_letter(col)
        ws.column_dimensions[col_letter].width = 20
        for row in range(1, ws.max_row + 1):  # 遍历所有行
            cell = ws[f"{col_letter}{row}"]
            cell.alignment = Alignment(horizontal='center', vertical='center')  # 设置单元格内容居中

    wb.save(file_path)

def is_same_with_last_row(ws, frame_result):
    """
    检查新行是否与最后一行相同"
    """
    last_row = ws.max_row
    last_row_odd_values = []
    # 从第三列开始，步长为 2
    for col in range(3, ws.max_column + 1, 2):
        cell = ws.cell(row=last_row, column=col)
        # 获取最后一行奇数列（除第一列）的数据，并将 None 转换为空字符串
        value = cell.value if cell.value is not None else ""
        last_row_odd_values.append(value)
    
    # 准备 frame_result 对应的比较值
    frame_result_values = []
    for name, data in frame_result.items():
        frame_result_values.append(data.get("result_text", ""))
    
    # 比较数据
    return last_row_odd_values == frame_result_values